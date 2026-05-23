# ==============================================================================
# AUTHOR METADATA & SUPPORT
# ==============================================================================
# Seller:        Jess Hayden
# Website:       jesshaydenconsulting.com
# Support Email: jlynne.hayden@gmail.com
#
# Need this pipeline integrated directly with your live clinical reporting API
# (Athenahealth, eCW, NextGen, etc.)? Reach out via my website or email to discuss
# bringing me on as a consultant to help with your data quality initiatives!
# ==============================================================================

# EHR COMPLIANCE & DATA QUALITY AUDIT REPORT

import pandas as pd
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ==========================================
# 1. LOAD AND AUDIT MESSY EHR DATA
# ==========================================
# Load the messy CSV data
df = pd.read_csv("messy_ehr_export.csv")

# Convert date columns to datetime objects for mathematical comparison
df['admission_date'] = pd.to_datetime(df['admission_date'], errors='coerce')
df['discharge_date'] = pd.to_datetime(df['discharge_date'], errors='coerce')
current_date = pd.Timestamp(datetime.date.today())

# Initialize lists to hold our audit tally findings
audit_tally = []

# --- Check A: Track Duplicate NPI Mappings Across Entire File ---
# Identity instances where an NPI maps to different doctor name variations
npi_groups = df.groupby('provider_npi')['provider_name'].nunique()
duplicate_npis = npi_groups[npi_groups > 1].index.tolist()

# --- ROW-BY-ROW VALIDATION ---
for idx, row in df.iterrows():
    patient = row['patient_id']
    
    # 1. Invalid/Missing NPI
    npi_raw = str(row['provider_npi']).strip().split('.')[0] # Clean float strings if any
    if pd.isna(row['provider_npi']) or len(npi_raw) != 10 or not npi_raw.isdigit():
        audit_tally.append({
            "Patient ID": patient, "Field": "provider_npi", 
            "Issue Type": "Invalid/Missing NPI", "Current Value": row['provider_npi'],
            "Action Required": "Verify provider NPI profile and update registry."
        })
        
    # 2. Timeline Discrepancy (Discharge before Admission)
    if pd.notna(row['admission_date']) and pd.notna(row['discharge_date']):
        if row['discharge_date'] < row['admission_date']:
            audit_tally.append({
                "Patient ID": patient, "Field": "discharge_date", 
                "Issue Type": "Timeline Discrepancy", 
                "Current Value": f"Adm: {row['admission_date'].strftime('%Y-%m-%d')} | Dis: {row['discharge_date'].strftime('%Y-%m-%d')}",
                "Action Required": "Review chart timeline; correct clinical dates."
            })

    # 3. Missing Billing Code
    if pd.isna(row['icd_10_code']) or str(row['icd_10_code']).strip() == "":
        audit_tally.append({
            "Patient ID": patient, "Field": "icd_10_code", 
            "Issue Type": "Missing Billing Code", "Current Value": "BLANK",
            "Action Required": "Route back to medical coding team for chart review."
        })

    # 4. Duplicate Provider Profile (New Check)
    if row['provider_npi'] in duplicate_npis:
        audit_tally.append({
            "Patient ID": patient, "Field": "provider_name",
            "Issue Type": "Duplicate Provider Profile", "Current Value": row['provider_name'],
            "Action Required": "Duplicate NPI mapping found. Merge clinician credentials in registry."
        })
        
    # 5. Future Encounter Date Error (New Check)
    if pd.notna(row['admission_date']) and row['admission_date'] > current_date:
        audit_tally.append({
            "Patient ID": patient, "Field": "admission_date",
            "Issue Type": "Future Encounter Date", "Current Value": row['admission_date'].strftime('%Y-%m-%d'),
            "Action Required": "Future date entry detected. Verify against physical intake sheets."
        })


# Convert audit findings into a clean dataframe
df_tally = pd.DataFrame(audit_tally)

# ==========================================
# 2. GENERATE POLISHED EXCEL WORKBOOK
# ==========================================
wb = Workbook()

# Styling Definitions
font_title = Font(name="Calibri", size=16, bold=True, color="1B365D")
font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
font_body = Font(name="Calibri", size=11, bold=False)
font_kpi_num = Font(name="Calibri", size=20, bold=True, color="1B365D")

fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
fill_zebra = PatternFill(start_color="F4F7FA", end_color="F4F7FA", fill_type="solid")
fill_error = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid") # Soft Red

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
)

# ------------------------------------------
# TAB 1: EXECUTIVE DASHBOARD
# ------------------------------------------
ws1 = wb.active
ws1.title = "Executive Summary"
ws1.views.sheetView[0].showGridLines = False # Remove default gridlines

# Title Block
ws1["A1"] = "EHR Compliance & Data Quality Executive Report"
ws1["A1"].font = font_title
ws1["A2"] = f"Report Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
ws1["A2"].font = Font(name="Calibri", size=10, italic=True, color="595959")

# Construct KPI Blocks
ws1["A4"] = "Total Records Audited"
ws1["A5"] = len(df)
ws1["B4"] = "Total Issues Detected"
ws1["B5"] = len(df_tally)

for col in ["A", "B"]:
    ws1[f"{col}4"].font = Font(name="Calibri", size=10, bold=True, color="595959")
    ws1[f"{col}5"].font = font_kpi_num
    ws1[f"{col}4"].alignment = Alignment(horizontal="center")
    ws1[f"{col}5"].alignment = Alignment(horizontal="center")
    ws1[f"{col}4"].fill = fill_zebra
    ws1[f"{col}5"].fill = fill_zebra

# Data Error Summary Table
ws1["A9"] = "Data Error Summary"
ws1["A9"].font = Font(name="Calibri", size=12, bold=True, color="1B365D")
ws1["A10"] = "Issue Type"
ws1["B10"] = "Count"
ws1["A10"].font = font_header
ws1["B10"].font = font_header
ws1["A10"].fill = fill_header
ws1["B10"].fill = fill_header

# Align headers nicely (left for text, right for numeric count)
ws1["A10"].alignment = Alignment(horizontal="left", vertical="center")
ws1["B10"].alignment = Alignment(horizontal="right", vertical="center")

# Define a thin grid border to replace the global gridlines inside this table
grid_side = Side(style='thin', color='D9D9D9') # Light gray color matching standard Excel gridlines
table_border = Border(left=grid_side, right=grid_side, top=grid_side, bottom=grid_side)

issue_counts = df_tally['Issue Type'].value_counts()
for idx, (issue, count) in enumerate(issue_counts.items(), start=11):
    ws1[f"A{idx}"] = issue
    ws1[f"B{idx}"] = count
    ws1[f"A{idx}"].font = font_body
    ws1[f"B{idx}"].font = font_body

    # Align the text to the left and row counts to the right
    ws1[f"A{idx}"].alignment = Alignment(horizontal="left", vertical="center")
    ws1[f"B{idx}"].alignment = Alignment(horizontal="right", vertical="center")

    # Apply the thin grid borders to just these cells 
    ws1[f"A{idx}"].border = table_border
    ws1[f"B{idx}"].border = table_border

    # Apply soft zebra striping for readability
    if idx % 2 == 0:
        ws1[f"A{idx}"].fill = fill_zebra
        ws1[f"B{idx}"].fill = fill_zebra

# Data Error Bar Chart
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

# 1. Initialize the Column Chart
chart = BarChart()
chart.type = "col"             # Vertical columns
chart.style = 10               # Clean dark-blue professional template
chart.title = "Distribution of Compliance Issues"
# chart.y_axis.title = "Count"
# chart.x_axis.title = "Issue Type"
chart.y_axis.majorGridlines = None # Remove horizontal gridlines for a cleaner look

# 2. Map the Data & Titles Natively (NO Loops to avoid XML corruption)
# data_ref points to the numbers including the "Count" header (Rows 10 to 13, Column B)
data_ref = Reference(ws1, min_col=2, min_row=10, max_row=10 + len(issue_counts))

# cats_ref points to the specific text strings (Rows 11 to 13, Column A)
cats_ref = Reference(ws1, min_col=1, min_row=11, max_row=10 + len(issue_counts))

# Add data directly. titles_from_data=True handles the internal naming mapping safely.
chart.add_data(data_ref, titles_from_data=True, from_rows=False)
chart.set_categories(cats_ref)

# 3. Clean up the Data Labels (Only show raw numbers on top of columns)
chart.dataLabels = DataLabelList()
chart.dataLabels.showVal = True       # Display number count cleanly
chart.dataLabels.showSerName = False  # Hide the default "Series 1" label which is redundant
chart.dataLabels.showCatName = False  # Keep messy floating text turned off

# 4. Enforce the Color Coding and Legend Layout
# Because openpyxl groups a single column dataset into one color by default,
# this command forces Excel to color each bar uniquely based on its text category!
chart.varyColors = True

chart.legend.position = "r"           # Place your pristine color legend on the right side
chart.width = 14                      
chart.height = 9.5

# 5. Place the Chart
ws1.add_chart(chart, "D1")

# ------------------------------------------
# TAB 2: THE ACTIONABLE TALLY SHEET
# ------------------------------------------
ws2 = wb.create_sheet(title="Data Quality Tally")
ws2.views.sheetView[0].showGridLines = True # Keep gridlines on the detailed sheet

# Append Headers to Tab 2
headers = list(df_tally.columns)
ws2.append(headers)

# Apply header styles
for col_num, header in enumerate(headers, 1):
    cell = ws2.cell(row=1, column=col_num)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal="left", vertical="center")

# Append row data and apply professional styling formatting
for row_idx, row_data in enumerate(dataframe_to_rows(df_tally, index=False, header=False), 2):
    ws2.append(row_data)
    for col_num in range(1, len(row_data) + 1):
        cell = ws2.cell(row=row_idx, column=col_num)
        cell.font = font_body
        cell.border = thin_border
        
        # Soft red fill for the specific error column to draw focus
        if col_num == 3: # Issue Type Column
            cell.fill = fill_error
        elif row_idx % 2 == 0: # Soft zebra striping for readability
            cell.fill = fill_zebra

# Freeze headers on data log so they scroll nicely
ws2.freeze_panes = "A2"

# Auto-fit columns to avoid truncated text or ### errors
for ws in [ws1, ws2]:
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

# Save the final automated file
wb.save("EHR_Compliance_Audit_Report.xlsx")
print("Process complete. Polished 5-point Audit Report generated successfully.")
